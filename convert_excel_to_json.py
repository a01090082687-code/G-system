"""
엑셀 파일을 읽어서 JavaScript 데이터 형식으로 변환하는 스크립트
인원명부(1,2공장 구분).xlsx 파일을 읽어서 전체 인원 데이터를 생성합니다.
"""

import json
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl이 설치되지 않았습니다. 설치 중...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

def read_excel_to_members(excel_path):
    """
    엑셀 파일을 읽어서 전체 인원 데이터를 생성합니다.
    미국/중국 주재원은 제외합니다.
    """
    try:
        wb = openpyxl.load_workbook(excel_path)
        members = []
        
        # 모든 시트 확인
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n시트 '{sheet_name}' 처리 중...")
            
            # 헤더 행 찾기 (이름, 부서, 공장 등의 컬럼)
            header_row = None
            name_col = None
            dept_col = None
            factory_col = None
            
            for row_idx, row in enumerate(sheet.iter_rows(max_row=10), start=1):
                row_values = [cell.value for cell in row]
                # 이름, 부서, 공장 등의 키워드로 헤더 찾기
                if any(keyword in str(val).lower() if val else False for val in row_values 
                       for keyword in ['이름', '성명', 'name', '부서', 'dept', '공장', 'factory']):
                    header_row = row_idx
                    for col_idx, val in enumerate(row_values, start=1):
                        val_str = str(val).lower() if val else ""
                        if '이름' in val_str or '성명' in val_str or 'name' in val_str:
                            name_col = col_idx
                        elif '부서' in val_str or 'dept' in val_str:
                            dept_col = col_idx
                        elif '공장' in val_str or 'factory' in val_str:
                            factory_col = col_idx
                    break
            
            if not header_row:
                print(f"  헤더를 찾을 수 없습니다. 첫 행을 헤더로 사용합니다.")
                header_row = 1
                name_col = 1
                dept_col = 2
                factory_col = 3
            
            # 데이터 읽기
            for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row + 1), start=header_row + 1):
                name = row[name_col - 1].value if name_col and name_col <= len(row) else None
                dept = row[dept_col - 1].value if dept_col and dept_col <= len(row) else None
                factory = row[factory_col - 1].value if factory_col and factory_col <= len(row) else None
                
                # 빈 행 스킵
                if not name or str(name).strip() == "":
                    continue
                
                name = str(name).strip()
                
                # 미국/중국 주재원 제외
                if any(keyword in name.lower() for keyword in ['미국', '중국', 'usa', 'china', 'us', 'cn']):
                    continue
                if dept and any(keyword in str(dept).lower() for keyword in ['미국', '중국', 'usa', 'china', 'us', 'cn']):
                    continue
                
                # 부서 정보 정리
                if not dept:
                    dept = ""
                else:
                    dept = str(dept).strip()
                
                # 공장 정보 추출 (부서에서 공장 정보 추출)
                factory_name = "1공장"
                if factory:
                    factory_str = str(factory).strip()
                    if "2" in factory_str or "2공장" in factory_str:
                        factory_name = "2공장"
                    elif "1" in factory_str or "1공장" in factory_str:
                        factory_name = "1공장"
                elif dept:
                    if "2공장" in dept or "2" in dept:
                        factory_name = "2공장"
                    elif "1공장" in dept or "1" in dept:
                        factory_name = "1공장"
                
                # 부서명 정리 (공장 정보 포함)
                if dept and factory_name not in dept:
                    dept = f"{factory_name} {dept}" if dept else factory_name
                elif not dept:
                    dept = factory_name
                
                members.append({
                    "name": name,
                    "dept": dept,
                    "factory": factory_name,
                    "checkupDate": None,
                    "status": "completed"  # 일반검진 완료
                })
                
                print(f"  추가: {name} ({dept})")
        
        return members
    
    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return []

def generate_js_data(members):
    """JavaScript 데이터 형식으로 변환"""
    js_code = f"""// 전체 인원명부 데이터 (자동 생성됨 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
// 엑셀 파일에서 자동으로 추출된 데이터입니다.

const ALL_MEMBERS_DATA = {json.dumps(members, ensure_ascii=False, indent=2)};

// 기존 APP_DATA.healthCheckupMembers에 병합
if (typeof APP_DATA !== 'undefined') {{
    APP_DATA.healthCheckupMembers = ALL_MEMBERS_DATA;
    console.log(`✅ 전체 인원 ${ALL_MEMBERS_DATA.length}명이 로드되었습니다.`);
}}
"""
    return js_code

if __name__ == "__main__":
    import os
    
    # 현재 스크립트 위치 기준으로 경로 설정
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(os.path.expanduser("~"), "Downloads", "인원명부(1,2공장 구분).xlsx")
    
    # 파일이 없으면 다른 경로 시도
    if not os.path.exists(excel_path):
        excel_path = os.path.join(script_dir, "..", "Downloads", "인원명부(1,2공장 구분).xlsx")
    
    print("=" * 60)
    print("엑셀 파일을 JavaScript 데이터로 변환합니다")
    print("=" * 60)
    print(f"\n파일 경로: {excel_path}")
    
    if not os.path.exists(excel_path):
        print(f"\n⚠️ 파일을 찾을 수 없습니다: {excel_path}")
        print("\n사용 방법:")
        print("1. 엑셀 파일 경로를 직접 입력하거나")
        print("2. 스크립트를 엑셀 파일과 같은 폴더에 두고 실행하세요.")
        excel_path = input("\n엑셀 파일 전체 경로를 입력하세요: ").strip().strip('"')
        if not os.path.exists(excel_path):
            print("파일을 찾을 수 없습니다.")
            sys.exit(1)
    
    members = read_excel_to_members(excel_path)
    
    if not members:
        print("\n⚠️ 인원 데이터를 찾을 수 없습니다.")
        print("엑셀 파일의 구조를 확인해주세요.")
        sys.exit(1)
    
    print(f"\n✅ 총 {len(members)}명의 인원이 추출되었습니다.")
    
    # JavaScript 파일로 저장
    output_path = os.path.join(script_dir, "js", "all-members-data.js")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    js_code = generate_js_data(members)
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(js_code)
    
    print(f"\n✅ JavaScript 파일이 생성되었습니다: {output_path}")
    
    # app.js에 직접 병합하는 코드 생성
    merge_code = f"""
// ===== 전체 인원명부 데이터 (엑셀에서 자동 생성) =====
// 생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
// 총 {len(members)}명

const ALL_MEMBERS_FROM_EXCEL = {json.dumps(members, ensure_ascii=False, indent=2)};

// APP_DATA.healthCheckupMembers에 병합
if (typeof APP_DATA !== 'undefined') {{
    APP_DATA.healthCheckupMembers = ALL_MEMBERS_FROM_EXCEL;
    console.log(`✅ 전체 인원 ${ALL_MEMBERS_FROM_EXCEL.length}명이 로드되었습니다.`);
    
    // 통계 업데이트
    if (typeof updateHealthCheckupStats === 'function') {{
        updateHealthCheckupStats();
    }}
    if (typeof renderHealthCheckupList === 'function') {{
        renderHealthCheckupList();
    }}
}}
"""
    
    merge_path = os.path.join(script_dir, "js", "merge-members.js")
    with open(merge_path, 'w', encoding='utf-8') as f:
        f.write(merge_code)
    
    print(f"\n✅ 병합 스크립트가 생성되었습니다: {merge_path}")
    print(f"\n다음 단계:")
    print(f"1. index.html의 </body> 태그 직전에 다음을 추가하세요:")
    print(f'   <script src="js/all-members-data.js"></script>')
    print(f"2. 또는 js/app.js 파일의 healthCheckupMembers 배열을 위 데이터로 교체하세요.")
    print(f"\n💡 빠른 병합: index.html에 <script src='js/merge-members.js'></script> 추가")
